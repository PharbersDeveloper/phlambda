from constants.Errors import Errors, FileNotFound, NotXlsxFile
from handler.Strategy import Strategy
from openpyxl.utils.exceptions import InvalidFileException
import os
import openpyxl
import constants.Common as Common
import logging


class Xlsx(Strategy):
    wb = None

    def __init__(self):
        pass

    def __init_openpyxl(self, path):
        self.wb = openpyxl.load_workbook(filename=path,
                                         read_only=True,
                                         keep_links=False,
                                         data_only=True)

    def __build_data(self, data):
        def filter_not_title(item):
            count = item[-1]
            is_none_content = item[-2]
            return count > 1 and is_none_content != "None"

        content = data  # list(filter(filter_not_title, data))
        read_num = (len(data) - len(content)) + 1
        sheet = content[0][0]
        schema = []
        for index, item in enumerate(content[0][1]):
            if item == "None":
                schema.append("col_{0}".format(index))
            else:
                schema.append(item)

        data = list(map(lambda x: x[1], content[1:]))
        return {
            "readNumber": read_num,
            "sheet": sheet,
            "schema": schema,
            "data": data
        }

    def __get_excel_data(self, sheets, out_num, skip_first, skip_next):
        def get_sheet(sheet):
            ws = self.wb[sheet]
            excel_data = []
            rows = ws.iter_rows(min_row=skip_first + 1 if skip_first > 0 else skip_first,
                                max_row=skip_first + out_num + 1)
            for row in rows:
                cells = [str(cell.value) for cell in row]
                none_set = "".join(list(set(cells)))
                excel_data.append((sheet, cells, none_set, len(none_set)))
            schema = excel_data.pop(0)
            excel_data = excel_data[skip_next:]
            excel_data.insert(0, schema)
            return excel_data

        data = list(map(get_sheet, sheets))
        self.wb.close()
        data = list(map(self.__build_data, data))
        return data

    def do_exec(self, data):
        logging.debug("Excel Xlsx === \n")
        logging.debug(data)
        result = {}
        try:
            skip_first = data["skip_first"]
            skip_next = data["skip_next"]
            project = data["project"]
            temp_file = data["tempfile"]
            sheet = data.get("sheet", "").strip()
            out_num = int(data.get("out_number", 0)) if int(data.get("out_number", 0)) > 0 else 20
            path = f"{os.environ.get(Common.PATH_PREFIX).format(project)}{temp_file}"

            self.__init_openpyxl(path)

            sheets = self.wb.sheetnames if not sheet else [sheet]
            result["sheets"] = self.wb.sheetnames
            result["body"] = self.__get_excel_data(sheets, out_num, skip_first, skip_next)
        except FileNotFoundError as e:
            raise FileNotFound(e)
        except InvalidFileException as e:
            raise NotXlsxFile(e)
        except Exception as e:
            raise Errors(e)

        return result
