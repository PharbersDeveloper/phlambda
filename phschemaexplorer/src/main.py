import openpyxl
import json
import os
import traceback
from time import perf_counter

__PATH_PREFIX = "PATH_PREFIX"


def __init_openpyxl(path):
    return openpyxl.load_workbook(filename=path, read_only=True, keep_links=False, data_only=True)


def get_excel_data(wb, sheets, out_number, skip_first, skip_next):
    def get_sheet(sheet):
        ws = wb[sheet]
        excel_data = []
        rows = ws.iter_rows(min_row=skip_first + 1 if skip_first == 0 else skip_first + 2 , max_row=skip_first + out_number + 1)
        for row in rows:
            cells = [str(cell.value) for cell in row]
            none_set = "".join(list(set(cells)))
            excel_data.append((sheet, cells, none_set, len(none_set)))
        schema = excel_data.pop(0)
        excel_data = excel_data[skip_next:]
        excel_data.insert(0, schema)
        return excel_data

    begin = perf_counter()
    data = list(map(get_sheet, sheets))
    wb.close()
    data = list(map(build_data, data))
    end = perf_counter()
    print("iterator {0:.2f}s".format(end - begin))
    begin = end
    return data


def build_data(data):
    def filter_not_title(item):
        count = item[-1]
        is_none_content = item[-2]
        return count > 1 and is_none_content != "None"

    begin = perf_counter()
    content = data # list(filter(filter_not_title, data))
    read_num = (len(data) - len(content)) + 1
    sheet = content[0][0]
    schema = []
    for index, item in enumerate(content[0][1]):
        if item == "None":
            schema.append("col_{0}".format(index))
        else:
            schema.append(item)

    data = list(map(lambda x: x[1], content[1:]))
    end = perf_counter()
    print("build data iterator {0:.2f}s".format(end - begin))
    begin = end
    return {
        "readNumber": read_num,
        "sheet": sheet,
        "schema": schema,
        "data": data
    }


def lambda_handler(event, context):
    try:
        body = json.loads(event["body"])
        skip_first = body.get("skip_first", 0)
        skip_next = body.get("skip_next", 0)
        path = os.environ.get(__PATH_PREFIX).format(project=body["project"]) + body["tempfile"]

        begin = perf_counter()
        wb = __init_openpyxl(path)
        end = perf_counter()
        print("open excel iterator {0:.2f}s".format(end - begin))
        begin = end

        sheets = wb.sheetnames if not body.get("sheet").strip() else [body.get("sheet")]
        out_number = int(body.get("out_number")) if int(body.get("out_number", 0)) > 0 else 20
        result = {
            "sheets": wb.sheetnames,
            "body": get_excel_data(wb, sheets, out_number, skip_first, skip_next)
        }

        return {
            "headers": {
                "Access-Control-Allow-Headers": "Content-Type",
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "OPTIONS,POST,GET,PATCH,DELETE",
            },
            "statusCode": 200,
            "body": json.dumps(result)
        }
    except Exception as e:
        print(e)
        traceback.print_exc()
        return {
            "headers": {
                "Access-Control-Allow-Headers": "Content-Type",
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "OPTIONS,POST,GET,PATCH,DELETE",
            },
            "statusCode": 500,
            "msg": e
        }
