#/usr/local/bin/python3
import openpyxl
import re
from math import floor
from time import perf_counter
from clickhouse_driver import Client
import uuid

class ExcelToClickHouse:
    def __init__(self,host,file_name,sheet_name,mapper,des_table_name,temp_dict,title_row=1,g_batch_size=10000):
        self.client = Client(host=host)
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.mapper = mapper
        self.table_name = des_table_name
        self.temp_dict = temp_dict
        self.title_row = int(title_row)
        self.g_batch_size = int(g_batch_size)
        self.batch_hit_time = 0

    #--计算dimensions
    def calCellsRange(self,dim):
        [left_top, right_bottom] = dim.split(':')
        left = re.findall(r'[A-Z]+', left_top)[0]
        top = int(re.findall(r'\d+', left_top)[0])
        right = re.findall(r'[A-Z]+', right_bottom)[0]
        bottom = int(re.findall(r'\d+', right_bottom)[0])
        return {'left': left, 'top': top, 'right': right, 'bottom': bottom}
    #--分批的索引
    def buildBatchIndex(self,data_rows_count,g_batch_size,title_row):
        batch_count = floor(data_rows_count / g_batch_size) + 1
        step_indeies = []
        for index in range(0, batch_count):
            step_indeies.append(
                min(index * g_batch_size + g_batch_size + title_row + 1, data_rows_count + title_row + 1))
        return step_indeies

    # header的索引，这里假设title只有一行
    def mapperAdapter(self,ws,dim,mapper, title_row, col_start, col_end):
        result = []
        header_cells = ws[dim['left'] + str(title_row) + ":" + dim['right'] + str(title_row)]
        for iter_mapper in mapper:
            for iter_header in header_cells[0]:
                if iter_mapper['src'] == iter_header.value:
                    result.append({'des': iter_mapper['des'], 'letter': iter_header.column_letter,
                                   'column': iter_header.column - 1})

        return result
    #--总行数
    def calDataRowsCount(self,dim,title_row):
        return dim['bottom'] - (title_row + 1) + 1

    def buildBatchCoordinate(self,dim, index_range):
        return dim['left'] + str(index_range.start) + ":" + dim['right'] + str(index_range.stop -1)
    def buildWriteSQL(self,ws,dim,mapper,title_row):
        adapted_mapper = self.mapperAdapter(ws, dim, mapper, title_row, col_start=None, col_end=None)
        static_col = ','.join(["measure","provider","version","owner"])
        cols_description = 'id,pkc,'+','.join(col['des'] for col in adapted_mapper) + ',dt,' + static_col
        sql = 'INSERT INTO ' + self.table_name + ' (' + cols_description + ') ' + 'VALUES'
        return sql,adapted_mapper

    def handle_stream_data(self):
        wb = openpyxl.load_workbook(filename=self.file_name, read_only=True, keep_links=False, data_only=True)
        ws = wb[self.sheet_name]
        dim = self.calCellsRange(ws.calculate_dimension())
        data_rows_count = self.calDataRowsCount(dim,self.title_row)
        step_indeies = self.buildBatchIndex(data_rows_count,self.g_batch_size,self.title_row)
        # 开始分配导入
        rows = ws.iter_rows(self.title_row + 1, max(step_indeies))
        row_process_count = self.title_row
        sql,adapted_mapper = self.buildWriteSQL(ws,dim,self.mapper,self.title_row)
        print(adapted_mapper)
        values = []
        begin = perf_counter()
        for row in rows:
            tmp = {}
            for col in adapted_mapper:
                tmp['id'] = str(uuid.uuid1())
                tmp[col['des']] = str(row[col['column']].value)
                tmp.update(self.temp_dict)
            values.append(tmp)
            row_process_count = row_process_count + 1
            if row_process_count == step_indeies[self.batch_hit_time] - 1:
                self.client.execute(sql, values)
                print(sql)
                print(values[0])
                print(len(values[0]))
                values.clear()
                self.batch_hit_time = self.batch_hit_time + 1
                end = perf_counter()
                print("iterator {0:.2f}s".format(end - begin))
                begin = end
        pass

def handle_mapper_args(mapper):
    mapper = list(map(lambda x: (list(x.values())[0],list(x.keys())[0]),mapper))
    mapper = [dict(zip(["src", "des"], i)) for i in mapper]
    return mapper

def lambda_handler(event,context):
    host='ec2-69-230-210-235.cn-northwest-1.compute.amazonaws.com.cn'
    #host = "192.168.0.66"
    event = event['parameters']['click_event']
    file_name = event["file_path"]
    sheet_name = event["sheet_name"]
    mapper = handle_mapper_args(event["mapper_args"])
    des_table_name = event["table_name"]
    title_row = 1 #event["begin_line"]
    g_batch_size = event["batch"]
    provider = event["provider"]
    version = event["version"]
    owner = event["owner"]
    dt = event["date"]
    static_key = ['pkc','measure','provider','version','owner','dt']
    static_value = ["","0",str(provider),str(version),str(owner),str(dt)]
    temp_dict = dict(zip(static_key,static_value))
    ExcelToClickHouse(host,file_name,sheet_name,mapper,des_table_name,temp_dict,title_row,g_batch_size).handle_stream_data()
    pass

