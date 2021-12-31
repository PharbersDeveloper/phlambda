# /usr/local/bin/python3
import re
import openpyxl
from math import floor


class Excel:

    def __init__(self, path, sheet_name, skip_first, skip_next, mapper, g_batch_size):
        self.wb = openpyxl.load_workbook(filename=path, read_only=True, keep_links=False, data_only=True)
        self.ws = self.wb[sheet_name]
        self.title_row = skip_first
        self.skip_next = skip_next
        self.mapper = mapper
        self.g_batch_size = g_batch_size
        self.dim = self.calCellsRange(self.ws.calculate_dimension())
        self.adapted_mapper = self.mapperAdapter(self.mapper, self.title_row, self.dim['left'], self.dim['right'])
        adapted_mapper = set(map(lambda item: item["des"], self.adapted_mapper))
        dim_mapper = set(map(lambda item: item["des"], self.mapper))
        # 这里判断，判断输入的参数是不是在你的数据dim之内
        print("Excel DIMS ===> \n")
        print(dim_mapper)
        print(adapted_mapper)

        if len(dim_mapper - adapted_mapper) != 0:
            raise Exception("DIM Don't match")
        self.data_rows_count = self.calDataRowsCount(self.dim)
        self.step_indeies = self.buildBatchIndex()

    # 计算dimensions
    def calCellsRange(self, dim):
        [left_top, right_bottom] = dim.split(':')
        left = re.findall(r'[A-Z]+', left_top)[0]
        top = int(re.findall(r'\d+', left_top)[0])
        right = re.findall(r'[A-Z]+', right_bottom)[0]
        bottom = int(re.findall(r'\d+', right_bottom)[0])
        return {'left': left, 'top': top, 'right': right, 'bottom': bottom}

    # 总行数
    def calDataRowsCount(self, dim):
        return dim['bottom'] - (self.title_row + 1) + 1

    # header的索引，这里假设title只有一行
    def mapperAdapter(self, mapper, title_row, col_start, col_end):
        result = []
        header_cells = self.ws[self.dim['left'] + str(title_row) + ":" + self.dim['right'] + str(title_row)]

        print("alex read header")
        header = list(map(lambda index, header_cell: {
            "value": header_cell.value if header_cell.value is not None else "col_{0}".format(index),
            "column_letter": header_cell.column_letter if header_cell.value is not None else None,
            "column": header_cell.column - 1 if header_cell.value is not None else index
        }, list(range(0, len(header_cells[0]) + 1)), header_cells[0]))

        for iter_mapper in mapper:
            for iter_header in header:
                if iter_mapper['src'] == iter_header["value"]:
                    result.append(
                        {'des': iter_mapper['des'],
                         'letter': iter_header["column_letter"],
                         'column': iter_header["column"]})

        # for iter_mapper in mapper:
        #     for iter_header in header_cells[0]:
        #         print(iter_header.value)
        #         if iter_mapper['src'] == iter_header.value:
        #             result.append(
        #                 {'des': iter_mapper['des'],
        #                  'letter': iter_header.column_letter,
        #                  'column': iter_header.column - 1})

        return result

    # 分批的索引
    def buildBatchIndex(self):
        batch_count = floor(self.data_rows_count / self.g_batch_size) + 1
        if self.data_rows_count == self.g_batch_size:
            batch_count = floor(self.data_rows_count / self.g_batch_size)
        step_indeies = []
        for index in range(0, batch_count):
            step_indeies.append(min(index * self.g_batch_size + self.g_batch_size + self.title_row + 1,
                                    self.data_rows_count + self.title_row + 1))
        return step_indeies

    def buildBatchCoordinate(self, dim, index_range):
        return dim['left'] + str(index_range.start) + ":" + dim['right'] + str(index_range.stop - 1)

    # 删除整行为None的数据
    def remove_none_value(self, lines):
        values = list(filter(lambda line: len(set(line.values())) != 1 or (len(set(line.values())) == 1 and set(line.values()).pop() is not None), lines))
        return values

    def replace_cell_none(self, lines):
        def conversion(line):
            keys = list(line.keys())
            values = list(line.values())
            update_index = list(map(lambda idx: idx[0], list(filter(lambda val: val[1] is None, list(enumerate(values))))))
            for index in update_index:
                values[index] = "None"
            return dict(zip(keys, values))

        return list(map(conversion, lines))

    def batchReader(self, func):
        # rows = self.ws.iter_rows(self.title_row + self.skip_next + 1 + 1, max(self.step_indeies)) # 多加一个1，因为是1-base
        rows = self.ws.iter_rows(self.title_row + 1 + self.skip_next, max(self.step_indeies))
        batch_hit_time = 0
        # row_process_count = self.title_row + self.skip_next + 1 + 1
        row_process_count = self.title_row + self.skip_next
        values = []
        for row in rows:
            tmp = {}
            for col in self.adapted_mapper:
                value = row[col['column']].value
                tmp[col['des']] = value
            values.append(tmp)
            row_process_count += 1
            if row_process_count == self.step_indeies[batch_hit_time] - 1:
                batch_hit_time = batch_hit_time + 1
                # 重构用
                # func(self.replace_cell_none(self.remove_none_value(values)), len(self.step_indeies), batch_hit_time)

                # 老方法
                func(self.replace_cell_none(self.remove_none_value(values)), self.adapted_mapper, len(self.step_indeies), batch_hit_time)
                values.clear()
                if batch_hit_time == len(self.step_indeies):
                    break

    @staticmethod
    def getSchema(path, sheet_name, skip_first):
        wb = openpyxl.load_workbook(filename=path, read_only=True, keep_links=False, data_only=True)
        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=skip_first, max_row=skip_first)
        cols = []
        for row in rows:
            for idx, cell in enumerate(row):
                if cell.value is None:
                    cell = "col_" + str(idx)
                    cols.append(cell)
                else:
                    cols.append(str(cell.value))
        return cols
